from boa_ultimate_service import ultimate, service
import openpyxl

def filter_rows_by_column_value(file_path, sheet_name, filter_column_index, filter_list):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)

    # Get the specified sheet
    sheet = wb[sheet_name]

    # Initialize an empty list to store the filtered rows
    filtered_rows = []

    # Iterate over the rows in the sheet
    for row in sheet.iter_rows(values_only=True):
        # Check if the value in the filter column matches any value in the filter list
        if row[filter_column_index - 1] in filter_list:  # Adjusting index to 0-based
            # Append the entire row to the filtered rows
            filtered_rows.append(row)

    return filtered_rows

if __name__ == "__main__":
    # Specify the file path of the Excel file
    excel_file_path = "resources/loads.xlsx"
    new_excel_file_path = "resources/filtered_loads.xlsx"

    # Specify the names of the sheets
    sheet_names = ["Joint Reactions"]

    # Define the filter lists for "Service" and "Ultimate"
    service_list = service
    ultimate_list = ultimate

    # Read the column names of the specified sheet to get the column index for filtering
    column_index = 4  # Assuming column 'D' is the fourth column (1-based index)

    # Filter rows based on column 'D' for "Service" and "Ultimate"
    service_rows = filter_rows_by_column_value(excel_file_path, sheet_names[0], column_index, service_list)
    ultimate_rows = filter_rows_by_column_value(excel_file_path, sheet_names[0], column_index, ultimate_list)

    # Create a new Excel workbook
    new_wb = openpyxl.Workbook()

    # Add "Service" and "Ultimate" sheets
    service_sheet = new_wb.create_sheet(title="Service")
    ultimate_sheet = new_wb.create_sheet(title="Ultimate")

    # Copy headers and rows 1 to 3 to the new sheets
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb[sheet_names[0]]
    for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
        service_sheet.append(row)
        ultimate_sheet.append(row)

    # Filter rows in "Service" sheet based on column 'D'
    for row in service_rows:
        service_sheet.append(row)

    # Filter rows in "Ultimate" sheet based on column 'D'
    for row in ultimate_rows:
        ultimate_sheet.append(row)

    # Save the new Excel workbook
    new_wb.save(new_excel_file_path)

    print("Filtered data has been saved to:", new_excel_file_path)



